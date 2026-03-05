#!/usr/bin/env python3
"""Generate Raito Business Overview Excel workbook."""
import sys
sys.path.insert(0, '/sessions/intelligent-wonderful-euler/mnt/dataset/scripts')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from parsers import consolidate_data
from config import (
    PRODUCT_NAMES, PRODUCT_SHORT, SELLING_PRICE_B2B, PRODUCTION_COST,
    MONTH_NAMES_HEB, TARGET_MONTHS_STOCK, PALLET_DIVISOR,
    extract_chain_name,
)

data = consolidate_data()
months = data['months']
products_all = ['chocolate', 'vanilla', 'mango', 'pistachio', 'dream_cake', 'magadat']
products_turbo = ['chocolate', 'vanilla', 'mango', 'pistachio']

wb = Workbook()

# ── Styles ──
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor='2C3E50')
sub_hdr_fill = PatternFill('solid', fgColor='34495E')
sub_hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
title_font = Font(name='Arial', bold=True, size=14, color='1E3A5F')
section_font = Font(name='Arial', bold=True, size=12, color='2563EB')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
total_fill = PatternFill('solid', fgColor='F0F0F0')
green_font = Font(name='Arial', bold=True, size=10, color='10B981')
red_font = Font(name='Arial', bold=True, size=10, color='EF4444')
orange_font = Font(name='Arial', bold=True, size=10, color='F59E0B')
blue_font = Font(name='Arial', bold=True, size=10, color='2563EB')
thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0'),
)
center = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
num_fmt = '#,##0'
currency_fmt = '₪#,##0'
pct_fmt = '0%'


def style_header_row(ws, row, cols, fill=None, font=None):
    f = fill or hdr_fill
    fn = font or hdr_font
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fn
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, cols, is_total=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold_font if is_total else normal_font
        cell.border = thin_border
        if is_total:
            cell.fill = total_fill
        if c > 1:
            cell.alignment = center


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════
# Sheet 1: Overview
# ═══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Overview'
ws.sheet_view.rightToLeft = False

r = 1
ws.cell(r, 1, 'Raito Business Overview').font = title_font
ws.cell(r, 3, f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M")}').font = Font(name='Arial', size=9, color='999999')
r += 2

# KPI Summary
ws.cell(r, 1, 'Monthly Sales Summary').font = section_font
r += 1
headers = ['Month', 'Total Units', 'Total Revenue (₪)', "Ma'ayan Units", 'Icedream Units', "Ma'ayan %", 'Icedream %']
for i, h in enumerate(headers, 1):
    ws.cell(r, i, h)
style_header_row(ws, r, len(headers))
r += 1

for mi, month in enumerate(months):
    md = data['monthly_data'][month]
    total_u = sum(md['combined'].get(p, {}).get('units', 0) for p in products_all)
    total_v = sum(md['combined'].get(p, {}).get('total_value', 0) for p in products_all)
    may_u = sum(md['mayyan'].get(p, {}).get('units', 0) for p in products_all)
    ice_u = sum(md['icedreams'].get(p, {}).get('units', 0) for p in products_all)
    ws.cell(r, 1, MONTH_NAMES_HEB.get(month, month)).font = bold_font
    ws.cell(r, 2, total_u).number_format = num_fmt
    ws.cell(r, 3, total_v).number_format = currency_fmt
    ws.cell(r, 4, may_u).number_format = num_fmt
    ws.cell(r, 5, ice_u).number_format = num_fmt
    ws.cell(r, 6).number_format = pct_fmt
    ws.cell(r, 6, may_u / total_u if total_u else 0)
    ws.cell(r, 7).number_format = pct_fmt
    ws.cell(r, 7, ice_u / total_u if total_u else 0)
    style_data_row(ws, r, len(headers))
    r += 1

# Total row
ws.cell(r, 1, 'Total').font = bold_font
for c in range(2, 8):
    col_letter = get_column_letter(c)
    first_data_row = r - len(months)
    last_data_row = r - 1
    if c <= 5:
        ws.cell(r, c, f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})')
        ws.cell(r, c).number_format = currency_fmt if c == 3 else num_fmt
    elif c == 6:
        ws.cell(r, c, f'=IF(B{r}>0,D{r}/B{r},0)')
        ws.cell(r, c).number_format = pct_fmt
    elif c == 7:
        ws.cell(r, c, f'=IF(B{r}>0,E{r}/B{r},0)')
        ws.cell(r, c).number_format = pct_fmt
style_data_row(ws, r, len(headers), is_total=True)
r += 2

set_col_widths(ws, [16, 14, 18, 16, 16, 13, 13])

# ═══════════════════════════════════════════════════════════════════════
# Sheet 2: Detailed Sales
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Detailed Sales')
ws2.sheet_view.rightToLeft = False
r = 1
ws2.cell(r, 1, 'Detailed Sales by Product & Month').font = section_font
r += 1

headers2 = ['Month', 'Product', "Ma'ayan (units)", 'Icedream (units)', 'Total Units', 'Revenue (₪)']
for i, h in enumerate(headers2, 1):
    ws2.cell(r, i, h)
style_header_row(ws2, r, len(headers2))
r += 1

for month in months:
    md = data['monthly_data'][month]
    mh = MONTH_NAMES_HEB.get(month, month)
    for p in products_all:
        c = md['combined'].get(p, {})
        u = c.get('units', 0)
        if u == 0:
            continue
        ws2.cell(r, 1, mh).font = bold_font
        ws2.cell(r, 2, PRODUCT_NAMES.get(p, p))
        ws2.cell(r, 3, c.get('mayyan_units', 0)).number_format = num_fmt
        ws2.cell(r, 4, c.get('icedreams_units', 0)).number_format = num_fmt
        ws2.cell(r, 5, u).number_format = num_fmt
        ws2.cell(r, 5).font = bold_font
        ws2.cell(r, 6, c.get('total_value', 0)).number_format = currency_fmt
        style_data_row(ws2, r, len(headers2))
        r += 1

set_col_widths(ws2, [14, 22, 16, 16, 14, 16])

# ═══════════════════════════════════════════════════════════════════════
# Sheet 3: Icedream Customers
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Icedream Customers')
ws3.sheet_view.rightToLeft = False
r = 1
ws3.cell(r, 1, 'Icedream Customers - By Product').font = section_font
r += 1

ice_pl = [p for p in ['chocolate', 'vanilla', 'mango', 'pistachio', 'magadat', 'dream_cake']]
ice_headers = ['Month', 'Customer']
for p in ice_pl:
    ice_headers.append(f'{PRODUCT_SHORT[p]} (units)')
for p in ice_pl:
    ice_headers.append(f'{PRODUCT_SHORT[p]} (₪)')
ice_headers += ['Total Units', 'Total ₪']

for i, h in enumerate(ice_headers, 1):
    ws3.cell(r, i, h)
style_header_row(ws3, r, len(ice_headers))
r += 1

for month in months:
    md = data['monthly_data'][month]
    mh = MONTH_NAMES_HEB.get(month, month)
    chains = {}
    for cust, pdata in md.get('icedreams_customers', {}).items():
        chain = extract_chain_name(cust)
        if chain not in chains:
            chains[chain] = {}
        for p in ice_pl:
            if p not in chains[chain]:
                chains[chain][p] = {'units': 0, 'value': 0}
            chains[chain][p]['units'] += pdata.get(p, {}).get('units', 0)
            chains[chain][p]['value'] += pdata.get(p, {}).get('value', 0)

    sorted_chains = sorted(chains.items(),
                           key=lambda x: sum(v.get('value', 0) for v in x[1].values()), reverse=True)
    for chain, pdata in sorted_chains:
        ctu = sum(pdata.get(p, {}).get('units', 0) for p in ice_pl)
        if ctu == 0:
            continue
        ctv = sum(pdata.get(p, {}).get('value', 0) for p in ice_pl)
        ws3.cell(r, 1, mh).font = bold_font
        ws3.cell(r, 2, chain).font = bold_font
        col = 3
        for p in ice_pl:
            u = pdata.get(p, {}).get('units', 0)
            ws3.cell(r, col, u if u else '').number_format = num_fmt
            col += 1
        for p in ice_pl:
            v = pdata.get(p, {}).get('value', 0)
            ws3.cell(r, col, v if v else '').number_format = currency_fmt
            col += 1
        ws3.cell(r, col, ctu).number_format = num_fmt
        ws3.cell(r, col).font = bold_font
        ws3.cell(r, col + 1, ctv).number_format = currency_fmt
        ws3.cell(r, col + 1).font = bold_font
        style_data_row(ws3, r, len(ice_headers))
        r += 1

widths3 = [14, 22] + [14] * (len(ice_pl) * 2) + [14, 14]
set_col_widths(ws3, widths3)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 4: Ma'ayan Chains
# ═══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Ma'ayan Chains")
ws4.sheet_view.rightToLeft = False
r = 1
ws4.cell(r, 1, "Ma'ayan Chains - By Product").font = section_font
r += 1

may_pl = ['chocolate', 'vanilla', 'mango', 'pistachio']
may_headers = ['Month', 'Chain']
for p in may_pl:
    may_headers.append(f'{PRODUCT_SHORT[p]} (units)')
for p in may_pl:
    may_headers.append(f'{PRODUCT_SHORT[p]} (₪)')
may_headers += ['Total Units', 'Total ₪']

for i, h in enumerate(may_headers, 1):
    ws4.cell(r, i, h)
style_header_row(ws4, r, len(may_headers))
r += 1

for month in months:
    md = data['monthly_data'][month]
    mh = MONTH_NAMES_HEB.get(month, month)
    cr = md.get('mayyan_accounts_revenue', {})
    norm_chains = {}
    for key, pdata in cr.items():
        source_chain, acct = key if isinstance(key, tuple) else ('', key)
        norm = extract_chain_name(acct, source_chain=source_chain)
        if norm not in norm_chains:
            norm_chains[norm] = {}
        for p in may_pl:
            if p not in norm_chains[norm]:
                norm_chains[norm][p] = {'units': 0, 'value': 0}
            pd_ = pdata.get(p, {})
            if isinstance(pd_, dict):
                norm_chains[norm][p]['units'] += pd_.get('units', 0)
                norm_chains[norm][p]['value'] += pd_.get('value', 0)

    sorted_chains = sorted(norm_chains.items(),
                           key=lambda x: sum(v.get('value', 0) for v in x[1].values()), reverse=True)
    for chain, pdata in sorted_chains:
        ctu = sum(pdata.get(p, {}).get('units', 0) for p in may_pl)
        if ctu == 0:
            continue
        ctv = sum(pdata.get(p, {}).get('value', 0) for p in may_pl)
        ws4.cell(r, 1, mh).font = bold_font
        ws4.cell(r, 2, chain).font = bold_font
        col = 3
        for p in may_pl:
            u = pdata.get(p, {}).get('units', 0)
            ws4.cell(r, col, u if u else '').number_format = num_fmt
            col += 1
        for p in may_pl:
            v = pdata.get(p, {}).get('value', 0)
            ws4.cell(r, col, round(v, 2) if v else '').number_format = currency_fmt
            col += 1
        ws4.cell(r, col, ctu).number_format = num_fmt
        ws4.cell(r, col).font = bold_font
        ws4.cell(r, col + 1, round(ctv, 2)).number_format = currency_fmt
        ws4.cell(r, col + 1).font = bold_font
        style_data_row(ws4, r, len(may_headers))
        r += 1

widths4 = [14, 28] + [14] * (len(may_pl) * 2) + [14, 14]
set_col_widths(ws4, widths4)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 5: Inventory
# ═══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Inventory')
ws5.sheet_view.rightToLeft = False

wh = data.get('warehouse', {})
dist_inv = data.get('dist_inv', {})
wh_products = wh.get('products', {})

r = 1
ws5.cell(r, 1, 'Total Available Stock — All Locations').font = section_font
r += 1

inv_headers = ['Product', 'Karfree (units)', 'Karfree (pallets)',
               'Icedream (units)', 'Icedream (pallets)',
               "Ma'ayan (units)", "Ma'ayan (pallets)",
               'Total Units', 'Total Pallets']
for i, h in enumerate(inv_headers, 1):
    ws5.cell(r, i, h)
style_header_row(ws5, r, len(inv_headers))
r += 1

prod_order = ['chocolate', 'vanilla', 'mango', 'pistachio', 'magadat', 'dream_cake']
for p in prod_order:
    wh_u = wh_products.get(p, {}).get('units', 0)
    ice_u = dist_inv.get('icedream', {}).get('products', {}).get(p, {}).get('units', 0)
    may_u = dist_inv.get('mayyan', {}).get('products', {}).get(p, {}).get('units', 0)
    total_u = wh_u + ice_u + may_u
    if total_u == 0:
        continue
    ws5.cell(r, 1, PRODUCT_NAMES.get(p, p)).font = bold_font
    ws5.cell(r, 2, wh_u).number_format = num_fmt
    ws5.cell(r, 3, round(wh_u / 2400, 1) if p != 'dream_cake' and wh_u else '-')
    ws5.cell(r, 4, ice_u).number_format = num_fmt
    ws5.cell(r, 5, round(ice_u / 2400, 1) if p != 'dream_cake' and ice_u else '-')
    ws5.cell(r, 6, may_u).number_format = num_fmt
    ws5.cell(r, 7, round(may_u / 2400, 1) if p != 'dream_cake' and may_u else '-')
    ws5.cell(r, 8, total_u).number_format = num_fmt
    ws5.cell(r, 8).font = bold_font
    ws5.cell(r, 9, round(total_u / 2400, 1) if p != 'dream_cake' else '-')
    style_data_row(ws5, r, len(inv_headers))
    r += 1

# Total row
wh_total = wh.get('total_units', 0)
ice_total = dist_inv.get('icedream', {}).get('total_units', 0)
may_total = dist_inv.get('mayyan', {}).get('total_units', 0)
grand_total = wh_total + ice_total + may_total
ws5.cell(r, 1, 'Total').font = bold_font
ws5.cell(r, 2, wh_total).number_format = num_fmt
ws5.cell(r, 3, round(wh_total / 2400, 1))
ws5.cell(r, 4, ice_total).number_format = num_fmt
ws5.cell(r, 5, round(ice_total / 2400, 1))
ws5.cell(r, 6, may_total).number_format = num_fmt
ws5.cell(r, 7, round(may_total / 2400, 1))
ws5.cell(r, 8, grand_total).number_format = num_fmt
ws5.cell(r, 9, round(grand_total / 2400, 1))
style_data_row(ws5, r, len(inv_headers), is_total=True)
r += 2

# ── Karfree Production Planning ──
ws5.cell(r, 1, 'Karfree Warehouse — Production Planning').font = section_font
r += 1
report_date = wh.get('report_date', 'N/A')
ws5.cell(r, 1, f'Report date: {report_date}').font = Font(name='Arial', size=9, italic=True, color='666666')
r += 1

kf_headers = ['Product', 'Units in Stock', 'Pallets', 'Share',
              'Earliest Expiry', 'Latest Expiry',
              'Avg Sales/Mo', 'Last Month Sales',
              'Months of Stock', 'Status', 'Suggested Prod.', 'Pallets Needed']
for i, h in enumerate(kf_headers, 1):
    ws5.cell(r, i, h)
style_header_row(ws5, r, len(kf_headers))
r += 1

total_units = wh.get('total_units', 0)
num_months_calc = len(months) if months else 1

for p in prod_order:
    pd = wh_products.get(p)
    if not pd:
        continue
    units = pd.get('units', 0)
    plt = pd.get('pallets', 0)
    pct = units / total_units if total_units > 0 else 0
    batches = pd.get('batches', [])
    expiry_dates = [b['expiry'] for b in batches if b.get('expiry')]
    earliest_exp = min(expiry_dates) if expiry_dates else '---'
    latest_exp = max(expiry_dates) if expiry_dates else '---'

    p_total = sum(data['monthly_data'].get(m, {}).get('combined', {}).get(p, {}).get('units', 0) for m in months)
    avg_monthly = round(p_total / num_months_calc) if num_months_calc > 0 else 0
    last_md = data['monthly_data'].get(months[-1], {}) if months else {}
    last_month_u = last_md.get('combined', {}).get(p, {}).get('units', 0)

    if avg_monthly > 0:
        months_stock = round(units / avg_monthly, 1)
    else:
        months_stock = 99 if units > 0 else 0

    if months_stock >= TARGET_MONTHS_STOCK * 1.5:
        status = 'OK'
        status_font = green_font
    elif months_stock >= TARGET_MONTHS_STOCK * 0.5:
        status = 'Low'
        status_font = orange_font
    else:
        status = 'Critical'
        status_font = red_font

    target_units = avg_monthly * TARGET_MONTHS_STOCK
    suggested = max(0, target_units - units)
    sug_pallets = '-' if p == 'dream_cake' else (round(suggested / PALLET_DIVISOR, 1) if suggested > 0 else '✓')

    ws5.cell(r, 1, PRODUCT_NAMES.get(p, p)).font = bold_font
    ws5.cell(r, 2, units).number_format = num_fmt
    ws5.cell(r, 3, plt)
    ws5.cell(r, 4, pct).number_format = pct_fmt
    ws5.cell(r, 5, earliest_exp)
    ws5.cell(r, 6, latest_exp)
    ws5.cell(r, 7, avg_monthly).number_format = num_fmt
    ws5.cell(r, 8, last_month_u).number_format = num_fmt
    ws5.cell(r, 9, months_stock)
    ws5.cell(r, 9).font = status_font
    ws5.cell(r, 10, status)
    ws5.cell(r, 10).font = status_font
    ws5.cell(r, 11, suggested if suggested > 0 else '✓').number_format = num_fmt
    ws5.cell(r, 12, sug_pallets)
    style_data_row(ws5, r, len(kf_headers))
    r += 1

set_col_widths(ws5, [22, 16, 10, 10, 16, 16, 16, 16, 16, 12, 18, 14])

# ═══════════════════════════════════════════════════════════════════════
# Sheet 6: Top Customers
# ═══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Top Customers')
ws6.sheet_view.rightToLeft = False
r = 1
ws6.cell(r, 1, 'Top Customers (All Months, by Units)').font = section_font
r += 1

all_c = {}
for month in months:
    md = data['monthly_data'][month]
    for c, pd in md.get('icedreams_customers', {}).items():
        chain = extract_chain_name(c)
        all_c[chain] = all_c.get(chain, 0) + sum(pd.get(p, {}).get('units', 0) for p in products_all)
    for key, pd in md.get('mayyan_accounts_revenue', {}).items():
        source_chain, acct = key if isinstance(key, tuple) else ('', key)
        norm = extract_chain_name(acct, source_chain=source_chain)
        k = f"Ma'ayan: {norm}"
        all_c[k] = all_c.get(k, 0) + sum(pd.get(p, {}).get('units', 0) for p in products_all if isinstance(pd.get(p), dict))

all_c = {k: v for k, v in all_c.items() if v > 0}
tc_list = sorted(all_c.items(), key=lambda x: x[1], reverse=True)[:20]

tc_headers = ['Rank', 'Customer', 'Total Units', 'Share']
for i, h in enumerate(tc_headers, 1):
    ws6.cell(r, i, h)
style_header_row(ws6, r, len(tc_headers))
r += 1

grand = sum(v for _, v in tc_list)
for rank, (name, units) in enumerate(tc_list, 1):
    ws6.cell(r, 1, rank).alignment = center
    ws6.cell(r, 2, name).font = bold_font
    ws6.cell(r, 3, units).number_format = num_fmt
    ws6.cell(r, 4, units / grand if grand else 0).number_format = pct_fmt
    style_data_row(ws6, r, len(tc_headers))
    r += 1

set_col_widths(ws6, [8, 30, 16, 12])

# ═══════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════
out_path = '/sessions/intelligent-wonderful-euler/mnt/dataset/docs/Raito_Business_Overview.xlsx'
wb.save(out_path)
print(f'Excel saved: {out_path}')
