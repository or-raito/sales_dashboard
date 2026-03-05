#!/usr/bin/env python3
"""
Raito Dashboard — Excel Report Generator
Generates comprehensive Excel workbook mirroring the dashboard.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import (
    fmt, fc, compute_kpis as _compute_kpis, count_pos as _count_pos,
    PRODUCT_NAMES, PRODUCT_SHORT, PRODUCT_STATUS, PRODUCT_COLORS,
    FLAVOR_COLORS, PRODUCTS_ORDER, SELLING_PRICE_B2B, PRODUCTION_COST,
    MONTH_NAMES_HEB, MONTH_ORDER, TARGET_MONTHS_STOCK, PALLET_DIVISOR,
    BRAND_FILTERS, CREATORS, DISTRIBUTOR_NAMES,
    OUTPUT_DIR, pallets,
)

def generate_excel(data):
    """Generate comprehensive Excel summary workbook mirroring the dashboard."""
    wb = Workbook()
    hf = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    hfill = PatternFill('solid', fgColor='1E3A5F')
    sub_hfill = PatternFill('solid', fgColor='3498DB')
    df_ = Font(name='Arial', size=10)
    nf = '#,##0'
    cf = '₪#,##0'
    cf2 = '₪#,##0.00'
    pf = '0.0%'
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    disc_f = Font(name='Arial', size=10, color='999999', italic=True)
    bold_f = Font(bold=True, name='Arial', size=10)
    title_f = Font(bold=True, size=14, name='Arial', color='1E3A5F')
    section_f = Font(bold=True, size=12, name='Arial', color='2C3E50')
    green_f = Font(bold=True, name='Arial', size=10, color='10B981')
    yellow_f = Font(bold=True, name='Arial', size=10, color='F59E0B')
    red_f = Font(bold=True, name='Arial', size=10, color='EF4444')
    green_fill = PatternFill('solid', fgColor='D1FAE5')
    yellow_fill = PatternFill('solid', fgColor='FEF3C7')
    red_fill = PatternFill('solid', fgColor='FEE2E2')

    months = data['months']
    num_months = len(months)
    products_order = ['chocolate', 'vanilla', 'mango', 'pistachio', 'dream_cake', 'magadat']
    wh = data.get('warehouse', {})
    wh_products = wh.get('products', {}) if wh else {}
    dist_inv = data.get('dist_inv', {})

    def _sh(ws, row, mc, fill=None):
        for c in range(1, mc+1):
            cell = ws.cell(row=row, column=c)
            cell.font = hf; cell.fill = fill or hfill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = tb

    def _sd(cell, disc=False):
        cell.font = disc_f if disc else df_
        cell.border = tb; cell.alignment = Alignment(horizontal='center')

    def _set_widths(ws, widths):
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w

    # ════════════════════════════════════════════════════════════════
    # Sheet 1: Overview (KPIs)
    # ════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Overview'
    ws1.cell(row=1, column=1, value='Raito Dashboard — Overview')
    ws1.cell(row=1, column=1).font = title_f

    # KPI cards
    r = 3
    ws1.cell(row=r, column=1, value='Period').font = section_f
    ws1.cell(row=r, column=2, value='Total Units').font = section_f
    ws1.cell(row=r, column=3, value='Total Revenue (₪)').font = section_f
    ws1.cell(row=r, column=4, value="Ma'ayan Units").font = section_f
    ws1.cell(row=r, column=5, value='Icedream Units').font = section_f
    ws1.cell(row=r, column=6, value="Ma'ayan %").font = section_f
    ws1.cell(row=r, column=7, value='Icedream %').font = section_f
    ws1.cell(row=r, column=8, value='POS Count').font = section_f
    _sh(ws1, r, 8)
    r = 4

    # Per-month KPIs
    kpi_start_row = r
    for month in months:
        tu, tr, tc, tgm, tmy, tic, mp, ip = _compute_kpis(data, [month])
        pos = _count_pos(data, [month])
        ws1.cell(row=r, column=1, value=MONTH_NAMES_HEB.get(month, month)); _sd(ws1.cell(row=r, column=1))
        ws1.cell(row=r, column=2, value=tu); _sd(ws1.cell(row=r, column=2)); ws1.cell(row=r, column=2).number_format = nf
        ws1.cell(row=r, column=3, value=round(tr)); _sd(ws1.cell(row=r, column=3)); ws1.cell(row=r, column=3).number_format = cf
        ws1.cell(row=r, column=4, value=tmy); _sd(ws1.cell(row=r, column=4)); ws1.cell(row=r, column=4).number_format = nf
        ws1.cell(row=r, column=5, value=tic); _sd(ws1.cell(row=r, column=5)); ws1.cell(row=r, column=5).number_format = nf
        ws1.cell(row=r, column=6, value=mp/100 if mp else 0); _sd(ws1.cell(row=r, column=6)); ws1.cell(row=r, column=6).number_format = pf
        ws1.cell(row=r, column=7, value=ip/100 if ip else 0); _sd(ws1.cell(row=r, column=7)); ws1.cell(row=r, column=7).number_format = pf
        ws1.cell(row=r, column=8, value=pos); _sd(ws1.cell(row=r, column=8)); ws1.cell(row=r, column=8).number_format = nf
        r += 1

    # Total row with formulas
    kpi_end_row = r - 1
    ws1.cell(row=r, column=1, value='Total'); ws1.cell(row=r, column=1).font = bold_f; ws1.cell(row=r, column=1).border = tb
    for col in [2, 3, 4, 5, 8]:
        cl = get_column_letter(col)
        ws1.cell(row=r, column=col).value = f'=SUM({cl}{kpi_start_row}:{cl}{kpi_end_row})'
        ws1.cell(row=r, column=col).font = bold_f; ws1.cell(row=r, column=col).border = tb
        ws1.cell(row=r, column=col).number_format = nf if col != 3 else cf
        ws1.cell(row=r, column=col).alignment = Alignment(horizontal='center')
    ws1.cell(row=r, column=6).value = f'=IF(B{r}>0,D{r}/B{r},0)'
    ws1.cell(row=r, column=6).font = bold_f; ws1.cell(row=r, column=6).border = tb
    ws1.cell(row=r, column=6).number_format = pf; ws1.cell(row=r, column=6).alignment = Alignment(horizontal='center')
    ws1.cell(row=r, column=7).value = f'=IF(B{r}>0,E{r}/B{r},0)'
    ws1.cell(row=r, column=7).font = bold_f; ws1.cell(row=r, column=7).border = tb
    ws1.cell(row=r, column=7).number_format = pf; ws1.cell(row=r, column=7).alignment = Alignment(horizontal='center')
    _set_widths(ws1, [16, 16, 20, 16, 16, 14, 14, 14])

    # ════════════════════════════════════════════════════════════════
    # Sheet 2: Sales by Flavor (units monthly breakdown)
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Sales by Flavor')
    ws2.cell(row=1, column=1, value='Units Sold by Flavor — Monthly Breakdown')
    ws2.cell(row=1, column=1).font = title_f

    headers2 = ['Product'] + [MONTH_NAMES_HEB.get(m, m) for m in months] + ['Total', 'Avg/Month', 'Share %']
    r = 3
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=r, column=c, value=h)
    _sh(ws2, r, len(headers2))
    r = 4
    flavor_start = r

    for p in products_order:
        has_data = False
        for month in months:
            u = data['monthly_data'].get(month, {}).get('combined', {}).get(p, {}).get('units', 0)
            if u > 0: has_data = True
        if not has_data:
            continue

        disc = PRODUCT_STATUS.get(p) == 'discontinued'
        ws2.cell(row=r, column=1, value=PRODUCT_NAMES.get(p, p))
        ws2.cell(row=r, column=1).font = disc_f if disc else bold_f
        ws2.cell(row=r, column=1).border = tb
        for i, month in enumerate(months):
            u = data['monthly_data'].get(month, {}).get('combined', {}).get(p, {}).get('units', 0)
            cell = ws2.cell(row=r, column=2+i, value=u if u else '')
            _sd(cell, disc); cell.number_format = nf

        # Total formula
        first_data_col = get_column_letter(2)
        last_data_col = get_column_letter(1 + num_months)
        total_col = 2 + num_months
        avg_col = total_col + 1
        share_col = avg_col + 1

        ws2.cell(row=r, column=total_col).value = f'=SUM({first_data_col}{r}:{last_data_col}{r})'
        ws2.cell(row=r, column=total_col).font = bold_f; ws2.cell(row=r, column=total_col).border = tb
        ws2.cell(row=r, column=total_col).number_format = nf; ws2.cell(row=r, column=total_col).alignment = Alignment(horizontal='center')

        tc_l = get_column_letter(total_col)
        ws2.cell(row=r, column=avg_col).value = f'=IF({num_months}>0,{tc_l}{r}/{num_months},0)'
        ws2.cell(row=r, column=avg_col).font = df_; ws2.cell(row=r, column=avg_col).border = tb
        ws2.cell(row=r, column=avg_col).number_format = nf; ws2.cell(row=r, column=avg_col).alignment = Alignment(horizontal='center')

        r += 1

    flavor_end = r - 1

    # Total row
    ws2.cell(row=r, column=1, value='Total'); ws2.cell(row=r, column=1).font = bold_f; ws2.cell(row=r, column=1).border = tb
    for c in range(2, share_col + 1):
        cl = get_column_letter(c)
        if c <= 1 + num_months or c == total_col or c == avg_col:
            ws2.cell(row=r, column=c).value = f'=SUM({cl}{flavor_start}:{cl}{flavor_end})'
        ws2.cell(row=r, column=c).font = bold_f; ws2.cell(row=r, column=c).border = tb
        ws2.cell(row=r, column=c).number_format = nf; ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    # Share % formulas (reference total row)
    tc_l = get_column_letter(total_col)
    sc_l = get_column_letter(share_col)
    for pr in range(flavor_start, flavor_end + 1):
        ws2.cell(row=pr, column=share_col).value = f'=IF({tc_l}{r}>0,{tc_l}{pr}/{tc_l}{r},0)'
        ws2.cell(row=pr, column=share_col).font = df_; ws2.cell(row=pr, column=share_col).border = tb
        ws2.cell(row=pr, column=share_col).number_format = pf; ws2.cell(row=pr, column=share_col).alignment = Alignment(horizontal='center')
    ws2.cell(row=r, column=share_col).value = f'=SUM({sc_l}{flavor_start}:{sc_l}{flavor_end})'
    ws2.cell(row=r, column=share_col).number_format = pf

    _set_widths(ws2, [22] + [14]*num_months + [14, 14, 12])

    # ── Distributor split sub-table ──
    r += 3
    ws2.cell(row=r, column=1, value="Sales Split by Distributor — Units")
    ws2.cell(row=r, column=1).font = section_f
    r += 1
    split_hdr = ['Product', 'Distributor'] + [MONTH_NAMES_HEB.get(m, m) for m in months] + ['Total']
    for c, h in enumerate(split_hdr, 1):
        ws2.cell(row=r, column=c, value=h)
    _sh(ws2, r, len(split_hdr))
    r += 1

    for p in products_order:
        has_data = any(data['monthly_data'].get(m, {}).get('combined', {}).get(p, {}).get('units', 0) > 0 for m in months)
        if not has_data:
            continue
        for dist_key, dist_label in [('mayyan_units', "Ma'ayan"), ('icedreams_units', 'Icedream')]:
            ws2.cell(row=r, column=1, value=PRODUCT_NAMES.get(p, p) if dist_key == 'mayyan_units' else '')
            ws2.cell(row=r, column=1).font = bold_f if dist_key == 'mayyan_units' else df_
            ws2.cell(row=r, column=1).border = tb
            ws2.cell(row=r, column=2, value=dist_label); _sd(ws2.cell(row=r, column=2))
            row_total = 0
            for i, month in enumerate(months):
                u = data['monthly_data'].get(month, {}).get('combined', {}).get(p, {}).get(dist_key, 0)
                cell = ws2.cell(row=r, column=3+i, value=u if u else '')
                _sd(cell); cell.number_format = nf
                row_total += u
            fc = get_column_letter(3)
            lc = get_column_letter(2 + num_months)
            cell = ws2.cell(row=r, column=3+num_months, value=f'=SUM({fc}{r}:{lc}{r})')
            cell.font = bold_f; cell.border = tb; cell.number_format = nf; cell.alignment = Alignment(horizontal='center')
            r += 1

    # ════════════════════════════════════════════════════════════════
    # Sheet 3: Inventory & Production Planning
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Inventory')
    ws3.cell(row=1, column=1, value='Inventory & Production Planning')
    ws3.cell(row=1, column=1).font = title_f

    # ── Total Available Stock ──
    r = 3
    ws3.cell(row=r, column=1, value='Total Available Stock — All Locations')
    ws3.cell(row=r, column=1).font = section_f
    r += 1
    inv_hdr = ['Product', 'Warehouse (Karfree)', 'Pallets', 'Icedream', 'Pallets', "Ma'ayan", 'Pallets', 'Total Available', 'Pallets']
    for c, h in enumerate(inv_hdr, 1):
        ws3.cell(row=r, column=c, value=h)
    _sh(ws3, r, len(inv_hdr))
    r += 1
    inv_start = r

    for p in products_order:
        wh_u = wh_products.get(p, {}).get('units', 0)
        ice_u = dist_inv.get('icedream', {}).get('products', {}).get(p, {}).get('units', 0)
        may_u = dist_inv.get('mayyan', {}).get('products', {}).get(p, {}).get('units', 0)
        total_u = wh_u + ice_u + may_u
        if total_u == 0:
            continue
        is_cake = p == 'dream_cake'

        ws3.cell(row=r, column=1, value=PRODUCT_NAMES.get(p, p)); ws3.cell(row=r, column=1).font = bold_f; ws3.cell(row=r, column=1).border = tb
        # Warehouse
        cell = ws3.cell(row=r, column=2, value=wh_u if wh_u else ''); _sd(cell); cell.number_format = nf
        cell = ws3.cell(row=r, column=3, value='-' if is_cake else (round(wh_u / 2400, 1) if wh_u else '')); _sd(cell)
        # Icedream
        cell = ws3.cell(row=r, column=4, value=ice_u if ice_u else ''); _sd(cell); cell.number_format = nf
        cell = ws3.cell(row=r, column=5, value='-' if is_cake else (round(ice_u / 2400, 1) if ice_u else '')); _sd(cell)
        # Ma'ayan
        cell = ws3.cell(row=r, column=6, value=may_u if may_u else ''); _sd(cell); cell.number_format = nf
        cell = ws3.cell(row=r, column=7, value='-' if is_cake else (round(may_u / 2400, 1) if may_u else '')); _sd(cell)
        # Total
        ws3.cell(row=r, column=8).value = f'=SUM(B{r},D{r},F{r})'
        ws3.cell(row=r, column=8).font = bold_f; ws3.cell(row=r, column=8).border = tb
        ws3.cell(row=r, column=8).number_format = nf; ws3.cell(row=r, column=8).alignment = Alignment(horizontal='center')
        cell = ws3.cell(row=r, column=9, value='-' if is_cake else (round(total_u / 2400, 1))); _sd(cell); cell.font = bold_f
        r += 1

    inv_end = r - 1
    # Total row
    ws3.cell(row=r, column=1, value='Total'); ws3.cell(row=r, column=1).font = bold_f; ws3.cell(row=r, column=1).border = tb
    for c in [2, 4, 6, 8]:
        cl = get_column_letter(c)
        ws3.cell(row=r, column=c).value = f'=SUM({cl}{inv_start}:{cl}{inv_end})'
        ws3.cell(row=r, column=c).font = bold_f; ws3.cell(row=r, column=c).border = tb
        ws3.cell(row=r, column=c).number_format = nf; ws3.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    for c in [3, 5, 7, 9]:
        cl_u = get_column_letter(c - 1)
        ws3.cell(row=r, column=c).value = f'=ROUND({cl_u}{r}/2400,1)'
        ws3.cell(row=r, column=c).font = bold_f; ws3.cell(row=r, column=c).border = tb; ws3.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    _set_widths(ws3, [22, 18, 10, 14, 10, 14, 10, 18, 10])

    # ── Karfree warehouse detail with expiry ──
    if wh_products:
        r += 3
        wh_date = wh.get('report_date', 'N/A')
        ws3.cell(row=r, column=1, value=f'Transfer Warehouse (Karfree) — as of {wh_date}')
        ws3.cell(row=r, column=1).font = section_f
        r += 1
        kh = ['Product', 'Units', 'Pallets', 'Share %', 'Earliest Expiry', 'Latest Expiry']
        for c, h in enumerate(kh, 1): ws3.cell(row=r, column=c, value=h)
        _sh(ws3, r, len(kh))
        r += 1
        wh_total = wh.get('total_units', 0)
        for p in ['chocolate', 'vanilla', 'mango', 'pistachio', 'dream_cake']:
            pd_ = wh_products.get(p)
            if not pd_: continue
            units = pd_.get('units', 0)
            batches = pd_.get('batches', [])
            expiry_dates = [b['expiry'] for b in batches if b.get('expiry')]
            ws3.cell(row=r, column=1, value=PRODUCT_NAMES.get(p, p)); ws3.cell(row=r, column=1).font = bold_f; ws3.cell(row=r, column=1).border = tb
            cell = ws3.cell(row=r, column=2, value=units); _sd(cell); cell.number_format = nf
            cell = ws3.cell(row=r, column=3, value=round(units / 2400, 1)); _sd(cell)
            cell = ws3.cell(row=r, column=4, value=round(units/wh_total, 3) if wh_total else 0); _sd(cell); cell.number_format = pf
            cell = ws3.cell(row=r, column=5, value=min(expiry_dates) if expiry_dates else '---'); _sd(cell)
            cell = ws3.cell(row=r, column=6, value=max(expiry_dates) if expiry_dates else '---'); _sd(cell)
            r += 1

    # ── Distributor stock detail ──
    for dist_key, dist_label in [('icedream', 'Icedream'), ('mayyan', "Ma'ayan")]:
        ddata = dist_inv.get(dist_key)
        if not ddata or not ddata.get('products'): continue
        r += 2
        d_date = ddata.get('report_date', 'N/A')
        ws3.cell(row=r, column=1, value=f'Distributor Inventory ({dist_label}) — as of {d_date}')
        ws3.cell(row=r, column=1).font = section_f
        r += 1
        dh = ['Product', 'Units', 'Pallets', 'Share %']
        for c, h in enumerate(dh, 1): ws3.cell(row=r, column=c, value=h)
        _sh(ws3, r, len(dh))
        r += 1
        d_total = ddata.get('total_units', 0)
        for p in products_order:
            pd_ = ddata['products'].get(p)
            if not pd_: continue
            units = pd_.get('units', 0)
            is_cake = p == 'dream_cake'
            ws3.cell(row=r, column=1, value=PRODUCT_NAMES.get(p, p)); ws3.cell(row=r, column=1).font = bold_f; ws3.cell(row=r, column=1).border = tb
            cell = ws3.cell(row=r, column=2, value=units); _sd(cell); cell.number_format = nf
            cell = ws3.cell(row=r, column=3, value='-' if is_cake else round(units / 2400, 1)); _sd(cell)
            cell = ws3.cell(row=r, column=4, value=round(units/d_total, 3) if d_total else 0); _sd(cell); cell.number_format = pf
            r += 1

    # ── Inventory Coverage & Production Planning ──
    r += 3
    ws3.cell(row=r, column=1, value=f'Inventory Coverage & Production Planning (target: {TARGET_MONTHS_STOCK} month stock)')
    ws3.cell(row=r, column=1).font = section_f
    r += 1
    cov_hdr = ['Product', 'Avg Sales/Mo', 'Last Month', 'Current Stock', 'Pallets', 'Months of Stock', 'Status', 'Suggested Production', 'Pallets']
    for c, h in enumerate(cov_hdr, 1): ws3.cell(row=r, column=c, value=h)
    _sh(ws3, r, len(cov_hdr))
    r += 1

    for p in products_order:
        p_total = sum(data['monthly_data'].get(m, {}).get('combined', {}).get(p, {}).get('units', 0) for m in months)
        avg_monthly = round(p_total / num_months) if num_months > 0 else 0
        last_md = data['monthly_data'].get(months[-1], {})
        last_month_u = last_md.get('combined', {}).get(p, {}).get('units', 0)
        wh_stock = wh_products.get(p, {}).get('units', 0)
        ice_stock = dist_inv.get('icedream', {}).get('products', {}).get(p, {}).get('units', 0)
        may_stock = dist_inv.get('mayyan', {}).get('products', {}).get(p, {}).get('units', 0)
        stock = wh_stock + ice_stock + may_stock

        if avg_monthly == 0 and stock == 0:
            continue

        months_stock = round(stock / avg_monthly, 1) if avg_monthly > 0 else (99 if stock > 0 else 0)
        target_units = avg_monthly * TARGET_MONTHS_STOCK
        suggested = max(0, target_units - stock)

        if months_stock >= TARGET_MONTHS_STOCK * 1.5:
            status_label = 'OK'; status_font = green_f; status_fill = green_fill
        elif months_stock >= TARGET_MONTHS_STOCK * 0.5:
            status_label = 'Low'; status_font = yellow_f; status_fill = yellow_fill
        else:
            status_label = 'Critical'; status_font = red_f; status_fill = red_fill

        is_cake = p == 'dream_cake'
        ws3.cell(row=r, column=1, value=PRODUCT_NAMES.get(p, p)); ws3.cell(row=r, column=1).font = bold_f; ws3.cell(row=r, column=1).border = tb
        cell = ws3.cell(row=r, column=2, value=avg_monthly); _sd(cell); cell.number_format = nf
        cell = ws3.cell(row=r, column=3, value=last_month_u); _sd(cell); cell.number_format = nf
        cell = ws3.cell(row=r, column=4, value=stock); _sd(cell); cell.number_format = nf; cell.font = bold_f
        cell = ws3.cell(row=r, column=5, value='-' if is_cake else round(stock / 2400, 1)); _sd(cell)
        cell = ws3.cell(row=r, column=6, value=months_stock); _sd(cell)
        cell = ws3.cell(row=r, column=7, value=status_label)
        cell.font = status_font; cell.fill = status_fill; cell.border = tb; cell.alignment = Alignment(horizontal='center')
        cell = ws3.cell(row=r, column=8, value=suggested if suggested > 0 else 0); _sd(cell); cell.number_format = nf; cell.font = bold_f
        cell = ws3.cell(row=r, column=9, value='-' if is_cake else (round(suggested / 2400, 1) if suggested > 0 else 0)); _sd(cell)
        r += 1

    # ════════════════════════════════════════════════════════════════
    # Sheet 4: Monthly Sales Detail (per-product per-month)
    # ════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Monthly Detail')
    ws4.cell(row=1, column=1, value='Monthly Sales — Product Detail')
    ws4.cell(row=1, column=1).font = title_f
    headers4 = ['Month', 'Product', 'Status', "Ma'ayan (units)", 'Icedream (units)', 'Total Units',
                "Ma'ayan Revenue (₪)", 'Icedream Revenue (₪)', 'Total Revenue (₪)']
    r = 3
    for c, h in enumerate(headers4, 1): ws4.cell(row=r, column=c, value=h)
    _sh(ws4, r, len(headers4))
    r += 1

    for month in months:
        md = data['monthly_data'][month]
        for p in data['products']:
            comb = md['combined'].get(p, {})
            units = comb.get('units', 0)
            if units == 0 and p in ['pistachio', 'magadat']:
                continue
            disc = PRODUCT_STATUS.get(p) == 'discontinued'
            vals = [MONTH_NAMES_HEB.get(month, month), PRODUCT_NAMES.get(p, p),
                    'Discontinued' if disc else ('New' if PRODUCT_STATUS.get(p) == 'new' else 'Active'),
                    comb.get('mayyan_units', 0), comb.get('icedreams_units', 0), units,
                    comb.get('mayyan_value', 0), comb.get('icedreams_value', 0), comb.get('total_value', 0)]
            for c, v in enumerate(vals, 1):
                cell = ws4.cell(row=r, column=c, value=v)
                _sd(cell, disc)
                if c in [4, 5, 6]: cell.number_format = nf
                elif c in [7, 8, 9]: cell.number_format = cf
            r += 1
    _set_widths(ws4, [16, 22, 14, 18, 18, 16, 20, 20, 20])

    # ════════════════════════════════════════════════════════════════
    # Sheet 5: Icedream Customers Detail
    # ════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('Icedream Customers')
    ws5.cell(row=1, column=1, value='Icedream — Customer & Product Detail')
    ws5.cell(row=1, column=1).font = title_f
    r = 3
    prod_list = ['chocolate', 'vanilla', 'mango', 'pistachio', 'magadat', 'dream_cake']
    for month in months:
        md = data['monthly_data'][month]
        custs = md.get('icedreams_customers', {})
        if not custs: continue
        custs_filtered = {c: pd for c, pd in custs.items()
                          if sum(v.get('units', 0) for v in pd.values()) > 0}
        if not custs_filtered: continue
        ws5.cell(row=r, column=1, value=MONTH_NAMES_HEB.get(month, month))
        ws5.cell(row=r, column=1).font = section_f
        r += 1
        ch = ['Customer'] + [PRODUCT_SHORT.get(p, p) + ' (units)' for p in prod_list] + ['Total Units', 'Total ₪']
        for c, h in enumerate(ch, 1): ws5.cell(row=r, column=c, value=h)
        _sh(ws5, r, len(ch)); r += 1
        cust_start = r
        for cust, pdata in custs_filtered.items():
            ws5.cell(row=r, column=1, value=cust); ws5.cell(row=r, column=1).font = df_; ws5.cell(row=r, column=1).border = tb
            tu = 0; tv = 0
            for i, p in enumerate(prod_list):
                u = pdata.get(p, {}).get('units', 0)
                v = pdata.get(p, {}).get('value', 0)
                tu += u; tv += v
                cell = ws5.cell(row=r, column=2+i, value=u if u else '')
                _sd(cell, PRODUCT_STATUS.get(p) == 'discontinued'); cell.number_format = nf
            total_col_u = 2 + len(prod_list)
            total_col_v = total_col_u + 1
            fc = get_column_letter(2); lc = get_column_letter(1 + len(prod_list))
            ws5.cell(row=r, column=total_col_u).value = f'=SUM({fc}{r}:{lc}{r})'
            ws5.cell(row=r, column=total_col_u).font = bold_f; ws5.cell(row=r, column=total_col_u).border = tb
            ws5.cell(row=r, column=total_col_u).number_format = nf; ws5.cell(row=r, column=total_col_u).alignment = Alignment(horizontal='center')
            cell = ws5.cell(row=r, column=total_col_v, value=round(tv, 2) if tv else '')
            _sd(cell); cell.number_format = cf; cell.font = bold_f
            r += 1
        r += 1
    for c in range(1, 2 + len(prod_list) + 3):
        ws5.column_dimensions[get_column_letter(c)].width = 14
    ws5.column_dimensions['A'].width = 22

    # ════════════════════════════════════════════════════════════════
    # Sheet 6: Ma'ayan Chains Detail
    # ════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Ma'ayan Chains")
    ws6.cell(row=1, column=1, value="Ma'ayan — Chain & Product Detail")
    ws6.cell(row=1, column=1).font = title_f
    r = 3
    ice_prods = ['chocolate', 'vanilla', 'mango']
    for month in months:
        md = data['monthly_data'][month]
        cr = md.get('mayyan_chains_revenue', {})
        if not cr: continue
        ws6.cell(row=r, column=1, value=MONTH_NAMES_HEB.get(month, month))
        ws6.cell(row=r, column=1).font = section_f
        r += 1
        ch = ['Chain'] + [PRODUCT_SHORT[p] + ' (units)' for p in ice_prods] + [PRODUCT_SHORT[p] + ' (₪)' for p in ice_prods] + ['Total Units', 'Total ₪']
        for c, h in enumerate(ch, 1): ws6.cell(row=r, column=c, value=h)
        _sh(ws6, r, len(ch)); r += 1
        sorted_chains = sorted(cr.items(), key=lambda x: sum(d.get('units', 0) for d in x[1].values()), reverse=True)
        for chain, pdata in sorted_chains:
            ws6.cell(row=r, column=1, value=chain); ws6.cell(row=r, column=1).font = df_; ws6.cell(row=r, column=1).border = tb
            tu = 0; tv = 0
            for i, p in enumerate(ice_prods):
                u = pdata.get(p, {}).get('units', 0) if isinstance(pdata.get(p), dict) else pdata.get(p, 0)
                v = pdata.get(p, {}).get('value', 0) if isinstance(pdata.get(p), dict) else round(u * SELLING_PRICE_B2B.get(p, 0), 2)
                tu += u; tv += v
                cell = ws6.cell(row=r, column=2+i, value=u if u else ''); _sd(cell); cell.number_format = nf
                cell = ws6.cell(row=r, column=2+len(ice_prods)+i, value=round(v, 2) if v else ''); _sd(cell); cell.number_format = cf
            cell = ws6.cell(row=r, column=2+2*len(ice_prods), value=tu); _sd(cell); cell.number_format = nf; cell.font = bold_f
            cell = ws6.cell(row=r, column=3+2*len(ice_prods), value=round(tv, 2)); _sd(cell); cell.number_format = cf; cell.font = bold_f
            r += 1
        r += 1
    for c in range(1, 10): ws6.column_dimensions[get_column_letter(c)].width = 16

    out = OUTPUT_DIR / 'supply_chain_summary.xlsx'
    wb.save(out)
    print(f"Excel saved: {out}")
    return out

