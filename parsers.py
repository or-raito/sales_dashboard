#!/usr/bin/env python3
"""
Raito Dashboard — Data Parsers
All file parsing functions: Icedream, Ma'ayan, Karfree, Distributor Stock, Production.
"""

import re
import math
from openpyxl import load_workbook
from config import (
    DATA_DIR, OUTPUT_DIR, classify_product, extract_units_per_carton,
    SELLING_PRICE_B2B, MONTH_ORDER,
)


# ── Month Detection ─────────────────────────────────────────────────────

def detect_month_from_sheet(ws):
    """Detect month from sheet name and date range row (Row 3)."""
    title = ws.title or ''
    month_keywords = {
        'דצמבר': 'December 2025', 'December': 'December 2025',
        'ינואר': 'January 2026', 'January': 'January 2026',
        'פברואר': 'February 2026', 'February': 'February 2026',
        'מרץ': 'March 2026', 'March': 'March 2026',
    }
    for keyword, month in month_keywords.items():
        if keyword in title:
            return month

    for col in range(1, 7):
        val = ws.cell(row=3, column=col).value
        if val is None:
            continue
        val = str(val)
        date_match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', val)
        if date_match:
            day, month_num, year = date_match.groups()
            month_num = int(month_num)
            year = int(year)
            month_map = {
                (12, 2025): 'December 2025',
                (1, 2026): 'January 2026', (2, 2026): 'February 2026',
                (3, 2026): 'March 2026', (4, 2026): 'April 2026',
            }
            return month_map.get((month_num, year), f'{month_num}/{year}')

    for row in range(1, 5):
        for col in range(1, 6):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            val = str(val)
            for keyword, month in month_keywords.items():
                if keyword in val:
                    return month
    return None


# ── Icedreams Parser ─────────────────────────────────────────────────────

def parse_icedreams_file(filepath):
    """Parse a single Icedreams monthly report."""
    wb = load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    month = detect_month_from_sheet(ws)
    if not month:
        month = ws.title if ws.title else 'Unknown'

    data = {'month': month, 'by_customer': {}, 'totals': {}}
    pending_products = []

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        is_bold = cell_a.font.bold if cell_a.font else False
        item_name = ws.cell(row=row_idx, column=4).value
        sales_val = ws.cell(row=row_idx, column=5).value
        quantity = ws.cell(row=row_idx, column=6).value

        if is_bold and cell_a.value:
            s = str(cell_a.value).strip()
            if 'סה"כ' in s and 'לדו' not in s and 'מס' not in s:
                cell_b = ws.cell(row=row_idx, column=2)
                customer_name = str(cell_b.value).strip() if cell_b.value else s.replace('סה"כ', '').strip()
                if customer_name:
                    if customer_name not in data['by_customer']:
                        data['by_customer'][customer_name] = {}
                    for pp in pending_products:
                        p, u, v, c = pp
                        if p not in data['by_customer'][customer_name]:
                            data['by_customer'][customer_name][p] = {'units': 0, 'value': 0, 'cartons': 0}
                        data['by_customer'][customer_name][p]['units'] += u
                        data['by_customer'][customer_name][p]['value'] += v
                        data['by_customer'][customer_name][p]['cartons'] += c
                    pending_products = []

        if item_name and quantity is not None:
            product = classify_product(item_name)
            if product:
                upc = extract_units_per_carton(item_name)
                raw_qty = float(quantity)
                # Negative = sales, Positive = returns (subtract)
                sign = -1 if raw_qty < 0 else 1  # sales are negative in report
                cartons = abs(raw_qty)
                units = round(cartons * upc) * sign * -1  # flip: negative→positive sales
                value = float(sales_val) * sign * -1 if sales_val else 0

                if product not in data['totals']:
                    data['totals'][product] = {'units': 0, 'value': 0, 'cartons': 0}
                data['totals'][product]['units'] += units
                data['totals'][product]['value'] += value
                data['totals'][product]['cartons'] += cartons * sign * -1
                pending_products.append((product, units, value, cartons * sign * -1))

    wb.close()
    return data


def parse_all_icedreams():
    """Parse all Icedreams files in the icedreams folder."""
    folder = DATA_DIR / 'icedreams'
    if not folder.exists():
        return {}
    results = {}
    for f in sorted(folder.glob('*.xlsx')):
        if f.name.startswith('~'):
            continue
        data = parse_icedreams_file(f)
        month = data['month']
        if month in results:
            for product, vals in data['totals'].items():
                if product not in results[month]['totals']:
                    results[month]['totals'][product] = {'units': 0, 'value': 0, 'cartons': 0}
                for k in ['units', 'value', 'cartons']:
                    results[month]['totals'][product][k] += vals[k]
            for cust, pdata in data.get('by_customer', {}).items():
                if cust not in results[month]['by_customer']:
                    results[month]['by_customer'][cust] = {}
                for p, vals in pdata.items():
                    if p not in results[month]['by_customer'][cust]:
                        results[month]['by_customer'][cust][p] = {'units': 0, 'value': 0, 'cartons': 0}
                    for k in ['units', 'value', 'cartons']:
                        results[month]['by_customer'][cust][p][k] += vals.get(k, 0)
        else:
            results[month] = data
    return results


# ── Ma'ayan Parser ───────────────────────────────────────────────────────

def parse_mayyan_file(filepath):
    """Parse Ma'ayan detailed distribution report."""
    import pandas as pd
    sheet_names = load_workbook(filepath, read_only=True).sheetnames

    detail_sheet = None
    for s in sheet_names:
        if 'פירוט' in s or 'דוח' in s:
            detail_sheet = s
            break
    if not detail_sheet:
        detail_sheet = sheet_names[-1] if len(sheet_names) > 1 else sheet_names[0]

    df = pd.read_excel(filepath, sheet_name=detail_sheet)

    month_col = next((c for c in df.columns if 'חודש' in str(c)), None)
    week_col = next((c for c in df.columns if 'שבועי' in str(c)), None)
    product_col = next((c for c in df.columns if 'פריט' in str(c)), None)
    units_col = next((c for c in df.columns if 'בודדים' in str(c)), None)
    chain_col = next((c for c in df.columns if 'רשת' in str(c)), None)
    type_col = next((c for c in df.columns if 'סוג' in str(c)), None)
    branch_col = next((c for c in df.columns if 'חשבון' in str(c) and 'שם' in str(c)), None)

    if not all([product_col, units_col]):
        return {}
    if not month_col and not week_col:
        return {}

    df['product'] = df[product_col].apply(classify_product)
    df = df[df['product'].notna()]

    if month_col:
        # Standard monthly format — map Hebrew month names to standard keys
        month_map = {}
        for m in df[month_col].unique():
            ms = str(m)
            if 'דצמבר' in ms or 'December' in ms:
                month_map[m] = 'December 2025'
            elif 'ינואר' in ms or 'January' in ms:
                month_map[m] = 'January 2026'
            elif 'פברואר' in ms or 'February' in ms:
                month_map[m] = 'February 2026'
            elif 'מרץ' in ms or 'March' in ms:
                month_map[m] = 'March 2026'
            else:
                month_map[m] = str(m)
        df['month_std'] = df[month_col].map(month_map)
    else:
        # Weekly format — no month column; infer month from filename or sheet names
        fname = str(filepath).lower()
        inferred_month = None
        month_keywords = {
            'dec': 'December 2025', 'דצמבר': 'December 2025',
            'jan': 'January 2026', 'ינואר': 'January 2026',
            'feb': 'February 2026', 'פברואר': 'February 2026',
            'mar': 'March 2026', 'מרץ': 'March 2026',
            'apr': 'April 2026', 'אפריל': 'April 2026',
        }
        for kw, month_val in month_keywords.items():
            if kw in fname:
                inferred_month = month_val
                break
        if not inferred_month:
            # Try sheet names
            for sn in sheet_names:
                for kw, month_val in month_keywords.items():
                    if kw in sn.lower():
                        inferred_month = month_val
                        break
                if inferred_month:
                    break
        if not inferred_month:
            inferred_month = 'Unknown'
        df['month_std'] = inferred_month
    results = {}

    for month in df['month_std'].unique():
        mdf = df[df['month_std'] == month]
        totals = {}
        for product in mdf['product'].unique():
            pdf = mdf[mdf['product'] == product]
            totals[product] = {
                'units': int(pdf[units_col].sum()),
                'value': 0,
                'transactions': int(len(pdf))
            }

        by_chain = {}
        if chain_col:
            for _, row in mdf.groupby([chain_col, 'product']).agg(
                total_units=(units_col, 'sum')
            ).reset_index().iterrows():
                chain = str(row[chain_col])
                if chain not in by_chain:
                    by_chain[chain] = {}
                by_chain[chain][row['product']] = int(row['total_units'])

        by_type = {}
        if type_col:
            for _, row in mdf.groupby([type_col, 'product']).agg(
                total_units=(units_col, 'sum')
            ).reset_index().iterrows():
                ct = str(row[type_col])
                if ct not in by_type:
                    by_type[ct] = {}
                by_type[ct][row['product']] = int(row['total_units'])

        by_account = {}
        if branch_col:
            group_cols = [branch_col, 'product']
            if chain_col:
                group_cols = [chain_col, branch_col, 'product']
            for _, row in mdf.groupby(group_cols).agg(
                total_units=(units_col, 'sum')
            ).reset_index().iterrows():
                acct = str(row[branch_col]).strip()
                chain_name = str(row[chain_col]).strip() if chain_col else ''
                key = (chain_name, acct)
                if key not in by_account:
                    by_account[key] = {}
                by_account[key][row['product']] = int(row['total_units'])

        branches = set()
        if branch_col:
            for b in mdf[branch_col].dropna().unique():
                branches.add(str(b).strip())

        results[month] = {
            'totals': totals,
            'by_chain': by_chain,
            'by_account': by_account,
            'by_customer_type': by_type,
            'branches': branches,
        }

    return results


def parse_all_mayyan():
    folder = DATA_DIR / 'mayyan'
    if not folder.exists():
        return {}
    results = {}
    for f in sorted(folder.glob('*.xlsx')):
        if f.name.startswith('~'):
            continue
        data = parse_mayyan_file(f)
        for month, mdata in data.items():
            if month not in results:
                results[month] = mdata
    return results


# ── Production Parser ─────────────────────────────────────────────────

def get_production_data():
    """Returns known production data. Will be extended to parse files."""
    production = {
    }
    folder = DATA_DIR / 'production'
    if folder.exists():
        for f in sorted(folder.glob('*.xlsx')):
            pass  # TODO: parse production files when format is known
    return production


# ── Karfree Warehouse Inventory Parser ────────────────────────────────

def _classify_product_karfree(text):
    """Classify product from reversed Hebrew PDF text."""
    t = text.lower()
    if 'וגנמ' in t or 'mango' in t:
        return 'mango'
    if 'דלוקוש' in t or 'chocolate' in t:
        return 'chocolate'
    if 'לינו' in t or 'vanilla' in t:
        return 'vanilla'
    if 'וקטסיפ' in t or 'pistachio' in t:
        return 'pistachio'
    if 'תגדגמ' in t or 'magadat' in t:
        return 'magadat'
    if 'תגוע' in t or 'dream' in t or 'cake' in t:
        return 'dream_cake'
    return None


def parse_karfree_inventory():
    """Parse Karfree cold storage PDF inventory reports."""
    folder = DATA_DIR / 'karfree'
    if not folder.exists():
        return {}

    pdf_files = sorted(folder.glob('*.pdf'))
    if not pdf_files:
        return {}

    try:
        import pdfplumber
    except ImportError:
        print("  Warning: pdfplumber not installed, skipping inventory")
        return {}

    filepath = pdf_files[-1]
    results = {
        'report_date': None,
        'products': {},
        'total_units': 0,
        'total_pallets': 0,
    }

    with pdfplumber.open(filepath) as pdf:
        full_text = ''
        for page in pdf.pages:
            full_text += page.extract_text() + '\n'

    lines = full_text.split('\n')
    current_product = None

    for line in lines:
        if ':ךיראתל ןוכנ' in line:
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            if date_match:
                results['report_date'] = date_match.group(1)

        if 'טירפ' in line and ':' in line:
            product = _classify_product_karfree(line)
            if product:
                current_product = product
                if product not in results['products']:
                    results['products'][product] = {
                        'units': 0, 'pallets': 0, 'batches': []
                    }

        if current_product and re.match(r'^\s*0\s+000017', line):
            parts = line.split()
            try:
                packages = None
                for i, p in enumerate(parts):
                    if p == '0.00' and i + 1 < len(parts):
                        packages = int(parts[i + 1])
                        break
                if packages:
                    factor = 10 if packages == 240 else 6
                    actual_units = packages * factor
                    dates = re.findall(r'\d{2}/\d{2}/\d{4}', line)
                    batch = {
                        'packages': packages,
                        'factor': factor,
                        'units': actual_units,
                        'expiry': dates[0] if len(dates) > 0 else None,
                        'production': dates[1] if len(dates) > 1 else None,
                        'entry': dates[2] if len(dates) > 2 else None,
                    }
                    results['products'][current_product]['batches'].append(batch)
            except (ValueError, IndexError):
                pass

        if 'טירפל' in line and 'כ"הס' in line:
            total_match = re.search(r'(\d[\d\s,]*\d)\s+(\d+)\s+:טירפל', line)
            if total_match and current_product:
                pallets_count = int(total_match.group(2))
                results['products'][current_product]['pallets'] = pallets_count

    for p, pdata in results['products'].items():
        pdata['units'] = sum(b['units'] for b in pdata['batches'])
        pdata['packages'] = sum(b['packages'] for b in pdata['batches'])
        results['total_units'] += pdata['units']
        results['total_pallets'] += pdata['pallets']

    return results


def get_warehouse_data():
    """Returns warehouse inventory data from Karfree reports."""
    return parse_karfree_inventory()


# ── Distributor Inventory Parser ──────────────────────────────────────

def parse_distributor_stock(filepath):
    """Parse a distributor stock Excel file (Icedream / Ma'ayan format)."""
    from pathlib import Path
    fp = Path(filepath)
    # Handle files without .xlsx extension
    if fp.suffix not in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        import shutil, tempfile
        tmp = Path(tempfile.mktemp(suffix='.xlsx'))
        shutil.copy2(fp, tmp)
        filepath = tmp
    wb = load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    results = {'products': {}, 'total_units': 0, 'report_date': None}

    for row_idx in range(1, min(5, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', str(val))
                if date_match:
                    results['report_date'] = date_match.group(1)
                    break
        if results['report_date']:
            break

    name_col = None
    qty_col = None
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            if 'שם פריט' in val or ('פריט' in val and 'מפתח' not in val and 'בר' not in val):
                name_col = col_idx
                header_row = row_idx
            if 'מלאי' in val or 'כמות' in val or 'יתרת' in val:
                qty_col = col_idx
        if name_col and qty_col:
            break

    if not name_col or not qty_col:
        wb.close()
        return results

    is_units_format = False
    start_row = (header_row + 1) if header_row else 5
    for test_row in range(start_row, min(start_row + 3, ws.max_row + 1)):
        test_name = str(ws.cell(row=test_row, column=name_col).value or '')
        if re.search(r'\d+/\d+', test_name) and not re.search(r'\*\s*\d+\s*יח', test_name):
            is_units_format = True
            break

    for row_idx in range(start_row, ws.max_row + 1):
        item_name = ws.cell(row=row_idx, column=name_col).value
        qty_val = ws.cell(row=row_idx, column=qty_col).value

        if not item_name or qty_val is None:
            continue
        item_name = str(item_name)

        if 'סה"כ' in item_name:
            continue

        product = classify_product(item_name)
        if not product:
            continue

        qty = float(qty_val)
        if is_units_format:
            units = max(0, int(qty))
            factor = 1
            cartons = qty
        else:
            factor = extract_units_per_carton(item_name)
            cartons = qty
            units = math.ceil(cartons * factor)

        if units <= 0:
            continue

        if product not in results['products']:
            results['products'][product] = {'units': 0, 'cartons': 0, 'factor': factor}
        results['products'][product]['units'] += units
        results['products'][product]['cartons'] += cartons
        results['total_units'] += units

    wb.close()
    return results


def get_distributor_inventory():
    """Parse distributor stock files from icedream and mayyan folders."""
    dist_inv = {}

    # Search in multiple possible locations for stock files
    ice_folders = [DATA_DIR / 'icedreams', OUTPUT_DIR / 'icedream']
    for icedream_folder in ice_folders:
        if not icedream_folder.exists():
            continue
        for f in sorted(icedream_folder.glob('*stock*'), key=lambda p: p.stat().st_mtime, reverse=True):
            if f.name.startswith('~'):
                continue
            try:
                data = parse_distributor_stock(f)
                if data.get('products'):
                    dist_inv['icedream'] = data
                    break
            except Exception:
                pass
        if 'icedream' in dist_inv:
            break

    may_folders = [DATA_DIR / 'mayyan', OUTPUT_DIR / 'Maayan']
    for mayyan_folder in may_folders:
        if not mayyan_folder.exists():
            continue
        for f in sorted(mayyan_folder.glob('*stock*'), key=lambda p: p.stat().st_mtime, reverse=True):
            if f.name.startswith('~'):
                continue
            try:
                data = parse_distributor_stock(f)
                if data.get('products'):
                    dist_inv['mayyan'] = data
                    break
            except Exception:
                pass
        if 'mayyan' in dist_inv:
            break

    return dist_inv


# ── Data Consolidation ──────────────────────────────────────────────────

def consolidate_data():
    """Merge all data sources into a unified dataset."""
    print("Processing Icedreams reports...")
    icedreams = parse_all_icedreams()
    print(f"  Found {len(icedreams)} months")

    print("Processing Ma'ayan reports...")
    mayyan = parse_all_mayyan()
    print(f"  Found {len(mayyan)} months")

    production = get_production_data()
    warehouse = get_warehouse_data()
    if warehouse:
        print(f"Processing Karfree inventory...")
        print(f"  Report date: {warehouse.get('report_date', 'N/A')}, Total: {warehouse.get('total_units', 0):,} units")

    dist_inv = get_distributor_inventory()
    for dist_name, dist_data in dist_inv.items():
        print(f"Processing {dist_name} stock...")
        print(f"  Report date: {dist_data.get('report_date', 'N/A')}, Total: {dist_data.get('total_units', 0):,} units")

    all_months = sorted(
        [m for m in set(list(icedreams.keys()) + list(mayyan.keys()))
         if m in MONTH_ORDER],
        key=lambda x: MONTH_ORDER.get(x, 99)
    )
    products = ['chocolate', 'vanilla', 'mango', 'magadat', 'dream_cake', 'pistachio']

    consolidated = {
        'months': all_months,
        'products': products,
        'monthly_data': {},
        'production': production,
        'warehouse': warehouse,
        'dist_inv': dist_inv,
    }

    for month in all_months:
        month_data = {
            'icedreams': icedreams.get(month, {}).get('totals', {}),
            'mayyan': mayyan.get(month, {}).get('totals', {}),
            'icedreams_customers': icedreams.get(month, {}).get('by_customer', {}),
            'mayyan_chains': mayyan.get(month, {}).get('by_chain', {}),
            'mayyan_accounts': mayyan.get(month, {}).get('by_account', {}),
            'mayyan_branches': mayyan.get(month, {}).get('branches', set()),
            'mayyan_types': mayyan.get(month, {}).get('by_customer_type', {}),
            'combined': {},
        }

        for p in products:
            ice_units = month_data['icedreams'].get(p, {}).get('units', 0)
            may_units = month_data['mayyan'].get(p, {}).get('units', 0)
            ice_value = month_data['icedreams'].get(p, {}).get('value', 0)
            may_value = round(may_units * SELLING_PRICE_B2B.get(p, 0), 2)
            total_value = round(ice_value + may_value, 2)
            from config import PRODUCTION_COST
            prod_cost_per_unit = PRODUCTION_COST.get(p, 0)
            total_units = ice_units + may_units
            total_prod_cost = round(total_units * prod_cost_per_unit, 2)
            gross_margin = round(total_value - total_prod_cost, 2) if p != 'magadat' else 0

            month_data['combined'][p] = {
                'units': total_units,
                'icedreams_units': ice_units,
                'mayyan_units': may_units,
                'icedreams_value': ice_value,
                'mayyan_value': may_value,
                'total_value': total_value,
                'production_cost': total_prod_cost,
                'gross_margin': gross_margin,
            }

        filtered_custs = {}
        for cust, pdata in month_data.get('icedreams_customers', {}).items():
            total_u = sum(v.get('units', 0) for v in pdata.values())
            if total_u > 0:
                for p, vals in pdata.items():
                    vals['value'] = round(vals.get('value', 0), 2)
                filtered_custs[cust] = pdata
        month_data['icedreams_customers'] = filtered_custs

        mayyan_chains_revenue = {}
        for chain, pdata in month_data.get('mayyan_chains', {}).items():
            mayyan_chains_revenue[chain] = {}
            for p, units in pdata.items():
                mayyan_chains_revenue[chain][p] = {
                    'units': units,
                    'value': round(units * SELLING_PRICE_B2B.get(p, 0), 2),
                }
        month_data['mayyan_chains_revenue'] = mayyan_chains_revenue

        mayyan_accounts_revenue = {}
        for key, pdata in month_data.get('mayyan_accounts', {}).items():
            # key is (chain_name, account_name) tuple
            mayyan_accounts_revenue[key] = {}
            for p, units in pdata.items():
                mayyan_accounts_revenue[key][p] = {
                    'units': units,
                    'value': round(units * SELLING_PRICE_B2B.get(p, 0), 2),
                }
        month_data['mayyan_accounts_revenue'] = mayyan_accounts_revenue

        consolidated['monthly_data'][month] = month_data

    return consolidated
